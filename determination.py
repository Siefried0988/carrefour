import openpyxl
from openpyxl.styles import Alignment
import time
import os
from openpyxl.styles import Font, colors
from export_log import exportLog

timestamp = time.strftime("%Y%m%d") 

excelName = 'Test Results('+str(timestamp)+').xlsx'
itemNumber = ['A','B','C','D']
#路徑為當前目錄的下一層叫做Test_Results的目錄
currentPath = os.path.dirname(os.path.realpath(__file__))
path = currentPath + '\\Test_Results'
#path = os.getcwd() + '\\Test_Results'
itemName = ['測項編號','測項名稱','測項結果','測試時間']


if not os.path.isdir(path):
    os.mkdir(path)


#將全部的字都靠左並且格子寬20個字元
def left(sheet):
    align = Alignment(horizontal='left')
    for numberOfRows in range(1, sheet.max_row + 1):
        for numberOfColumns in range(len(itemName)):
            sheet[itemNumber[numberOfColumns] + str(numberOfRows)].alignment = align
            #設定每格寬20個字元
            #分別是ABCD行
            sheet.column_dimensions[itemNumber[numberOfColumns]].width = 20.0
          

def changeColor(sheet, testResults):
    #從C2開始是因為C1是放標題
    for numberOfRows in range(2, sheet.max_row + 1):
        sheet.title = 'Font'
        passColor = Font(color = '228B22') #綠色 這串數字為色號
        failColor = Font(color = colors.RED) # 紅色
        
        #C行負責顯示Pass或是Fail
        if sheet['C' + str(numberOfRows)].value == 'Pass':
            sheet['C' + str(numberOfRows)].font = passColor
        else:
            sheet['C' + str(numberOfRows)].font = failColor


def exportResults(testResults, logName):
    #如果excel已經存在，就打開excel並且編輯
    if excelName in os.listdir(path):
        wb = openpyxl.load_workbook(path + '\\' + excelName)
        sheet = wb.active
        
        sheet.append(testResults)
        
        left(sheet)
        changeColor(sheet, testResults)
        wb.save(path + '\\' + excelName)
        exportLog('已經新增資料', logName)
    #如果excel還不存在就新增一個excel並且編輯
    else:  
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(itemName)
        sheet.title='Font'
        Bold = Font(bold=True)
        #將A1、B1、C1、D1都改成粗體字
        for numberOfRows in range(len(itemName)):
            sheet[itemNumber[numberOfRows] + '1'].font = Bold
        sheet.append(testResults)
        
        left(sheet)
        changeColor(sheet, testResults)
        wb.save(path + '\\' + excelName)
        exportLog('已新增excel及資料', logName)





